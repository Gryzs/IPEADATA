import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font

headers = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36"
}

urls = [
    ('http://www.ipeadata.gov.br/ExibeSerie.aspx?stub=1&serid=36482&module=M', 'IPCA'),
    ('http://www.ipeadata.gov.br/ExibeSerie.aspx?stub=1&serid=37796&module=M', 'IGP-M'),
    ('http://www.ipeadata.gov.br/ExibeSerie.aspx?stub=1&serid=32237&module=M', 'CDI'),
    ('http://www.ipeadata.gov.br/ExibeSerie.aspx?stub=1&serid=36472&module=M', 'INPC'),
    ('http://www.ipeadata.gov.br/ExibeSerie.aspx?stub=1&serid=38590&module=M', 'CÂMBIO'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=1678536198&MINDATA=1990&MAXDATA=2025&TNIVID=2&TPAID=1&module=R', 'POPULAÇÃO'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=1540855420&MINDATA=2014&MAXDATA=2025&TNIVID=2&TPAID=1&module=R', 'PIB ESTADUAL'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=2058319060&MINDATA=2023&MAXDATA=2025&TNIVID=2&TPAID=1&module=R', 'EMPREGADOS ADMISSÕES'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=2058319061&MINDATA=2023&MAXDATA=2025&TNIVID=2&TPAID=1&module=R', 'EMPREGADOS DEMISSÕES'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=1828975897&MINDATA=2023&MAXDATA=2025&TNIVID=2&TPAID=1&module=R', 'FOB'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=2096726935&MINDATA=2012&MAXDATA=2025&TNIVID=0&TPAID=1&module=S', 'GINI'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=40037&MINDATA=2012&MAXDATA=2025&TNIVID=0&TPAID=1&module=S', 'IDHM'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=2096726928&MINDATA=2022&MAXDATA=2025&TNIVID=0&TPAID=1&module=S', 'TAXA DESEMPREGO'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=2096726934&MINDATA=2012&MAXDATA=2025&TNIVID=0&TPAID=1&module=S', 'TAXA POBREZA'),
    ('http://www.ipeadata.gov.br/ExibeSerieR.aspx?stub=1&serid=2096726779&MINDATA=2023&MAXDATA=2025&TNIVID=0&TPAID=1&module=S', 'BOLSA FAMILIA')
]

# Criar um novo arquivo Excel
wb = Workbook()
wb.remove(wb.active)  # Remove a planilha padrão

for url, sheet_name in urls:
    site = requests.get(url, headers=headers)
    status = site.status_code

    if status == 200:
        soup = BeautifulSoup(site.content, 'html.parser')
        table = soup.find('table', id='grd_DXMainTable')

        if table:
            data = []
            rows = table.find_all('tr')

            # Armazenar a primeira linha em uma variável
            header_row = [cell.get_text(strip=True) for cell in rows[0].find_all('td')]
            ws = wb.create_sheet(title=sheet_name)

            # Remover espaços e duplicatas do cabeçalho
            seen = set()
            header_row = [x for x in header_row if x and not (x in seen or seen.add(x))]

            # Adicionar o cabeçalho à planilha e aplicar negrito
            ws.append(header_row)
            for cell in ws[1]:  # Acessa a primeira linha
                cell.font = Font(bold=True)  # Define o texto como negrito

            for row in rows[1:]:
                if row.get('id') == 'grd_DXHeadersRow0':
                    continue
                
                cells = row.find_all('td')
                row_data = [cell.get_text(strip=True) for cell in cells]

                if any(row_data) and not all(cell == '' for cell in row_data):
                    data.append(row_data)

            for row in data:
                if row and row[0] not in header_row:
                    ws.append(row)

            print(f"Dados extraídos e adicionados à planilha '{sheet_name}'.")
        else:
            print(f"Tabela não encontrada no link: {url}")
    else:
        print(f"Erro ao acessar o site. Status code: {status}")

# Salvar o arquivo Excel
excel_file = 'dados_tabelas.xlsx'
wb.save(excel_file)
print(f"Dados salvos com sucesso em '{excel_file}'.")